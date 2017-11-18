from openpyxl import *

class ExcelWriter:
    def __init__(self):
        print("Initializng Excel Writer")

    def ExcelTest(self):
        multiArray=[[1,2,3,4],[1,2,'a','b'], [3,4,'c','d'], [6,7,'e','f'],[2,1,'a','b'], [4,3,'c','d'], [7,6,'e','f']] 
        wb=load_workbook('test.xlsx')
        ws= wb.active
        mArrayLen=len(multiArray)
        for i in range(0,mArrayLen):
            ws.append(multiArray[i])
        # for i in range(4,mArrayLen):
        #     for j in range(0,len(multiArray[i])):
        #         ws.cell(row=i+2,column=j+1).value=multiArray[i][j]
        wb.save('test.xlsx')

    def InitExcelSheet(self):
        InitString=['Title','Cost','SKU','Page Yield','Color', 'Product Type','Compatible Printers']
        wb=Workbook()
        ws=wb.active
        ws.append(InitString)
        wb.save('TonerLandExcel.xlsx')

    def appendToExcelSheet(self,ExcelArray):
        wb=load_workbook('TonerLandExcel.xlsx')
        ws=wb.active
        ws.append(ExcelArray)
        wb.save('TonerLandExcel.xlsx')


