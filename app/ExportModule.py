from openpyxl import Workbook
import random

workbook = Workbook()

def TestCreate():
    headers = ['This', 'Right', 'Here', 'Is', 'My', 'Swag']
    excelSheet = workbook.create_sheet("Test_Sheet")
    # Set up headers
    for i in range(0, len(headers)):
        cell = excelSheet.cell(row = 1, column= i+1)
        cell.value = headers[i]
    
    for col in range(1, len(headers) + 1):
        for row in range(2, 10):
            cell = excelSheet.cell(row = row, column = col)
            cell.value = random.randint(1, 69)
    workbook.save("test.xlsx")

def create_excel_sheet_from_data(excelObjList, name):
    excelObj = excelObjList[0]
    excelSheet = workbook.create_sheet('Results')
    properties = []
    for prop in vars(excelObj).items():
        properties.append(prop[0].upper())
    print(properties)
    # Setup headers
    propertyPosition = []
    for i in range(0, len(properties)):
        index = i + 1
        propertyPosition.append((properties[i], index))
        cell = excelSheet.cell(row = 1, column = i + 1)
        cell.value = properties[i]
    
    for row in range(0, len(excelObjList)):
        for col in range(0, len(propertyPosition)):
            cell = excelSheet.cell(row = row + 2, column = col + 1)
            cell.value = getValue(excelObjList[row], propertyPosition[col])
    workbook.save(name+'.xlsx')
    return 'ok'

def getValue(excelObj, propertyInfo):
    propertyField = propertyInfo[0]
    if propertyField == 'URL':
        return excelObj.url
    elif propertyField == 'PRICE':
        return excelObj.price
    elif propertyField == 'VENDOR':
        return excelObj.vendor
    elif propertyField == 'NAME':
        return excelObj.name
    else:
        return ''
